from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.utils import timezone
from django.db.models import Q
from django.core.files.base import ContentFile
import json
import pandas as pd
import os
import io
from datetime import datetime, timedelta

from .models import DocumentTemplate, GeneratedDocument, TaskInstance, ThermometerVerificationRecord, TemperatureLog
from .document_template_serializers import DocumentTemplateSerializer, GeneratedDocumentSerializer
from .permissions import IsManagerForWriteOrAuthenticatedReadOnly

class DocumentTemplateViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing document templates.
    Only managers can create, update, or delete templates in their department.
    """
    serializer_class = DocumentTemplateSerializer
    permission_classes = [IsManagerForWriteOrAuthenticatedReadOnly]
    
    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return DocumentTemplate.objects.none()
            
        # Filter by template_type if provided
        template_type = self.request.query_params.get('template_type', None)
        queryset = DocumentTemplate.objects.all()
        
        if template_type:
            queryset = queryset.filter(template_type=template_type)
            
        # Superusers can see all templates
        if user.is_superuser:
            return queryset
        
        # Managers and staff can only see templates for their department
        try:
            if user.profile.department:
                return queryset.filter(department=user.profile.department)
            else:
                return DocumentTemplate.objects.none()
        except:
            return DocumentTemplate.objects.none()
    
    def perform_create(self, serializer):
        # Set created_by to the requesting user
        serializer.save(created_by=self.request.user)
        
    @action(detail=False, methods=['get'], url_path='by-department/(?P<department_id>[^/.]+)')
    def by_department(self, request, department_id=None):
        """
        Returns templates for a specific department.
        """
        templates = self.get_queryset().filter(department_id=department_id)
        serializer = self.get_serializer(templates, many=True)
        return Response(serializer.data)
        
    @action(detail=False, methods=['get'], url_path='types')
    def template_types(self, request):
        """
        Returns the available template types.
        """
        return Response(dict(DocumentTemplate.TEMPLATE_TYPE_CHOICES))
        
    @action(detail=True, methods=['post'], url_path='preview')
    def preview_document(self, request, pk=None):
        """
        Generates a preview of the document with validation results.
        """
        template = self.get_object()
        parameters = request.data.get('parameters', {})
        
        # Validate parameters
        validation_results = self._validate_parameters(template, parameters)
        
        # Generate preview data based on template type
        preview_data = self._generate_preview_data(template, parameters)
        
        return Response({
            'preview_data': preview_data,
            'validation_results': validation_results
        })
    
    def _validate_parameters(self, template, parameters):
        """
        Validates the parameters for document generation.
        """
        validation_results = {
            'is_valid': True,
            'has_critical_issues': False,
            'checks': {}
        }
        
        # Check date range
        start_date = parameters.get('startDate')
        end_date = parameters.get('endDate')
        
        if not start_date or not end_date:
            validation_results['checks']['date_range'] = {
                'passed': False,
                'message': 'Start date and end date are required.',
                'is_critical': True
            }
            validation_results['is_valid'] = False
            validation_results['has_critical_issues'] = True
        else:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                
                if end_date < start_date:
                    validation_results['checks']['date_range'] = {
                        'passed': False,
                        'message': 'End date cannot be before start date.',
                        'is_critical': True
                    }
                    validation_results['is_valid'] = False
                    validation_results['has_critical_issues'] = True
                elif (end_date - start_date).days > 90:
                    validation_results['checks']['date_range'] = {
                        'passed': False,
                        'message': 'Date range exceeds 90 days. This may result in a very large document.',
                        'is_critical': False
                    }
                    validation_results['is_valid'] = False
                else:
                    validation_results['checks']['date_range'] = {
                        'passed': True,
                        'message': f'Date range is valid: {start_date} to {end_date}',
                        'is_critical': False
                    }
            except ValueError:
                validation_results['checks']['date_range'] = {
                    'passed': False,
                    'message': 'Invalid date format. Use YYYY-MM-DD.',
                    'is_critical': True
                }
                validation_results['is_valid'] = False
                validation_results['has_critical_issues'] = True
        
        # Check data availability based on template type
        if template.template_type == 'temperature':
            has_data = self._check_temperature_data_availability(template, parameters)
            validation_results['checks']['temperature_data'] = {
                'passed': has_data,
                'message': 'Temperature data is available.' if has_data else 'No temperature data found for the selected period.',
                'is_critical': False
            }
            if not has_data:
                validation_results['is_valid'] = False
        
        elif template.template_type == 'cleaning':
            has_data = self._check_cleaning_data_availability(template, parameters)
            validation_results['checks']['cleaning_data'] = {
                'passed': has_data,
                'message': 'Cleaning task data is available.' if has_data else 'No cleaning task data found for the selected period.',
                'is_critical': False
            }
            if not has_data:
                validation_results['is_valid'] = False
        
        elif template.template_type == 'verification':
            has_data = self._check_verification_data_availability(template, parameters)
            validation_results['checks']['verification_data'] = {
                'passed': has_data,
                'message': 'Thermometer verification data is available.' if has_data else 'No thermometer verification data found for the selected period.',
                'is_critical': False
            }
            if not has_data:
                validation_results['is_valid'] = False
        
        return validation_results
    
    def _check_temperature_data_availability(self, template, parameters):
        """
        Check if temperature data is available for the given parameters.
        """
        try:
            start_date = datetime.strptime(parameters.get('startDate'), '%Y-%m-%d').date()
            end_date = datetime.strptime(parameters.get('endDate'), '%Y-%m-%d').date()
            
            # Check if there are any temperature logs in the date range
            count = TemperatureLog.objects.filter(
                recorded_at__date__gte=start_date,
                recorded_at__date__lte=end_date,
                thermometer__department=template.department
            ).count()
            
            return count > 0
        except:
            return False
    
    def _check_cleaning_data_availability(self, template, parameters):
        """
        Check if cleaning task data is available for the given parameters.
        """
        try:
            start_date = datetime.strptime(parameters.get('startDate'), '%Y-%m-%d').date()
            end_date = datetime.strptime(parameters.get('endDate'), '%Y-%m-%d').date()
            
            # Check if there are any task instances in the date range
            count = TaskInstance.objects.filter(
                scheduled_date__gte=start_date,
                scheduled_date__lte=end_date,
                cleaning_item__department=template.department
            ).count()
            
            return count > 0
        except:
            return False
    
    def _check_verification_data_availability(self, template, parameters):
        """
        Check if thermometer verification data is available for the given parameters.
        """
        try:
            start_date = datetime.strptime(parameters.get('startDate'), '%Y-%m-%d').date()
            end_date = datetime.strptime(parameters.get('endDate'), '%Y-%m-%d').date()
            
            # Check if there are any verification records in the date range
            count = ThermometerVerificationRecord.objects.filter(
                date_verified__gte=start_date,
                date_verified__lte=end_date,
                thermometer__department=template.department
            ).count()
            
            return count > 0
        except:
            return False
    
    def _generate_preview_data(self, template, parameters):
        """
        Generate preview data based on template type and parameters.
        """
        preview_data = {
            'sections': []
        }
        
        try:
            start_date = datetime.strptime(parameters.get('startDate'), '%Y-%m-%d').date()
            end_date = datetime.strptime(parameters.get('endDate'), '%Y-%m-%d').date()
            
            if template.template_type == 'temperature' and parameters.get('includeTemperatureLogs', True):
                # Get temperature logs for preview
                logs = TemperatureLog.objects.filter(
                    recorded_at__date__gte=start_date,
                    recorded_at__date__lte=end_date,
                    thermometer__department=template.department
                ).order_by('-recorded_at')[:20]  # Limit to 20 for preview
                
                if logs.exists():
                    log_data = []
                    for log in logs:
                        log_data.append({
                            'Date': log.recorded_at.strftime('%Y-%m-%d'),
                            'Time': log.recorded_at.strftime('%H:%M'),
                            'Thermometer': log.thermometer.name,
                            'Temperature': f"{log.temperature}°C",
                            'Recorded By': log.recorded_by.username,
                            'Notes': log.notes or ''
                        })
                    
                    preview_data['sections'].append({
                        'title': 'Temperature Logs',
                        'columns': ['Date', 'Time', 'Thermometer', 'Temperature', 'Recorded By', 'Notes'],
                        'data': log_data
                    })
            
            if template.template_type == 'cleaning' and parameters.get('includeCleaningTasks', True):
                # Get cleaning tasks for preview
                tasks = TaskInstance.objects.filter(
                    scheduled_date__gte=start_date,
                    scheduled_date__lte=end_date,
                    cleaning_item__department=template.department
                ).order_by('-scheduled_date')[:20]  # Limit to 20 for preview
                
                if tasks.exists():
                    task_data = []
                    for task in tasks:
                        task_data.append({
                            'Date': task.scheduled_date.strftime('%Y-%m-%d'),
                            'Item': task.cleaning_item.name,
                            'Status': task.get_status_display(),
                            'Completed By': task.completed_by.username if task.completed_by else '',
                            'Completed At': task.completed_at.strftime('%Y-%m-%d %H:%M') if task.completed_at else '',
                            'Notes': task.notes or ''
                        })
                    
                    preview_data['sections'].append({
                        'title': 'Cleaning Tasks',
                        'columns': ['Date', 'Item', 'Status', 'Completed By', 'Completed At', 'Notes'],
                        'data': task_data
                    })
            
            if template.template_type == 'verification' and parameters.get('includeThermometerVerifications', True):
                # Get verification records for preview
                records = ThermometerVerificationRecord.objects.filter(
                    date_verified__gte=start_date,
                    date_verified__lte=end_date,
                    thermometer__department=template.department
                ).order_by('-date_verified')[:20]  # Limit to 20 for preview
                
                if records.exists():
                    record_data = []
                    for record in records:
                        record_data.append({
                            'Date': record.date_verified.strftime('%Y-%m-%d'),
                            'Thermometer': record.thermometer.serial_number,
                            'Model': getattr(record.thermometer, 'model', 'N/A'),
                            'Calibrated Instrument': record.calibrated_instrument_no,
                            'Reading': f"{record.reading_after_verification}°C",
                            'Result': 'Passed',
                            'Verified By': record.calibrated_by.username if record.calibrated_by else 'Unknown',
                            'Corrective Action': record.corrective_action or 'None',
                            'Created At': record.created_at.strftime('%Y-%m-%d %H:%M') if record.created_at else 'N/A'
                        })
                    
                    preview_data['sections'].append({
                        'title': 'Thermometer Verifications',
                        'columns': ['Date', 'Thermometer', 'Result', 'Verified By', 'Notes'],
                        'data': record_data
                    })
        except Exception as e:
            # Add an error section if something goes wrong
            preview_data['sections'].append({
                'title': 'Error Generating Preview',
                'columns': ['Error'],
                'data': [{'Error': str(e)}]
            })
        
        return preview_data


def generate_document_file(template, parameters, user):
    """
    Generate a document file based on the template and parameters.
    Returns a tuple of (file_content, filename, error_message).
    """
    try:
        # Get the template file
        template_file_path = template.template_file.path
        
        # Parse parameters
        start_date = datetime.strptime(parameters.get('startDate'), '%Y-%m-%d').date()
        end_date = datetime.strptime(parameters.get('endDate'), '%Y-%m-%d').date()
        
        # Create a new workbook based on the template
        try:
            # Load the template workbook using openpyxl directly for more control
            from openpyxl import load_workbook
            wb = load_workbook(template_file_path)
            
            # Get verification records
            if template.template_type == 'verification' and parameters.get('includeThermometerVerifications', True):
                records = ThermometerVerificationRecord.objects.filter(
                    date_verified__gte=start_date,
                    date_verified__lte=end_date,
                    thermometer__department=template.department
                ).order_by('-date_verified')
                
                if records.exists():
                    # Get the active worksheet (usually the first one)
                    ws = wb.active
                    
                    # Add date range to the worksheet (typically in cell A1 or a header area)
                    date_range_text = f"Date Range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
                    if 'A1' in ws and not ws['A1'].value:
                        ws['A1'] = date_range_text
                    
                    # Find the starting row for data (typically after headers)
                    # This is a simplified approach - in a real implementation, you would
                    # look for specific markers or headers in the template
                    start_row = 5  # Assuming data starts at row 5
                    
                    # Populate the verification data
                    for i, record in enumerate(records):
                        row = start_row + i
                        
                        # Map data to columns - adjust column letters based on your template
                        ws.cell(row=row, column=1, value=record.date_verified.strftime('%Y-%m-%d'))  # Date
                        ws.cell(row=row, column=2, value=record.thermometer.serial_number)  # Thermometer
                        ws.cell(row=row, column=3, value=record.calibrated_instrument_no)  # Calibrated Instrument
                        ws.cell(row=row, column=4, value=f"{record.reading_after_verification}°C")  # Reading
                        ws.cell(row=row, column=5, value=record.calibrated_by.username if record.calibrated_by else "Unknown")  # Verified By
                        ws.cell(row=row, column=6, value=record.corrective_action or "None")  # Corrective Action
            
            # Save the workbook to a BytesIO object
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Get the file content
            file_content = output.getvalue()
            
            # Generate a filename
            filename = f"{template.name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
            
            return file_content, filename, None
            
        except Exception as e:
            return None, None, f"Error processing Excel template: {str(e)}"
        
    except Exception as e:
        return None, None, f"Error generating document: {str(e)}"


class GeneratedDocumentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing generated documents.
    Only managers can create documents in their department.
    """
    serializer_class = GeneratedDocumentSerializer
    permission_classes = [IsManagerForWriteOrAuthenticatedReadOnly]
    
    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return GeneratedDocument.objects.none()
            
        # Filter by status if provided
        status_filter = self.request.query_params.get('status', None)
        queryset = GeneratedDocument.objects.all()
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
            
        # Superusers can see all generated documents
        if user.is_superuser:
            return queryset
        
        # Managers and staff can only see documents for their department
        try:
            if user.profile.department:
                return queryset.filter(department=user.profile.department)
            else:
                return GeneratedDocument.objects.none()
        except:
            return GeneratedDocument.objects.none()
    
    def create(self, request, *args, **kwargs):
        # Extract data from request
        template_id = request.data.get('template_id')
        department_id = request.data.get('department_id')
        parameters = request.data.get('parameters', {})
        
        # Validate required fields
        if not template_id or not department_id or not parameters:
            return Response(
                {"detail": "Missing required fields: template_id, department_id, or parameters"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Get the template
            template = DocumentTemplate.objects.get(id=template_id)
            
            # Generate the document file
            file_content, filename, error_message = generate_document_file(template, parameters, request.user)
            
            if error_message:
                # Create a failed document record
                document = GeneratedDocument.objects.create(
                    template=template,
                    department_id=department_id,
                    generated_by=request.user,
                    status='failed',
                    error_message=error_message,
                    parameters=parameters
                )
                
                return Response(
                    {"detail": error_message, "document_id": document.id},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Create the document file
            document = GeneratedDocument(
                template=template,
                department_id=department_id,
                generated_by=request.user,
                status='completed',
                parameters=parameters
            )
            
            # Save the generated file
            document.generated_file.save(filename, ContentFile(file_content))
            document.save()
            
            # Serialize and return the document
            serializer = self.get_serializer(document)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except DocumentTemplate.DoesNotExist:
            return Response(
                {"detail": "Template not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"detail": f"Error generating document: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def perform_create(self, serializer):
        # Set generated_by to the requesting user
        serializer.save(generated_by=self.request.user)
        
    @action(detail=False, methods=['get'], url_path='recent')
    def recent_documents(self, request):
        """
        Returns the most recent generated documents for the user's department.
        """
        documents = self.get_queryset().order_by('-created_at')[:10]
        serializer = self.get_serializer(documents, many=True)
        return Response(serializer.data)
        
    @action(detail=False, methods=['get'], url_path='by-template/(?P<template_id>[^/.]+)')
    def by_template(self, request, template_id=None):
        """
        Returns documents generated from a specific template.
        """
        documents = self.get_queryset().filter(template_id=template_id)
        serializer = self.get_serializer(documents, many=True)
        return Response(serializer.data)
